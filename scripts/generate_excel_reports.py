"""
Shared Excel report generator, reused across all four real test suites.
Reads the same JSON each suite already produces (backend_security_results.json,
api_load_test_results.json, web_e2e_results.json, mobile_e2e_results.json) —
no changes to the suites themselves. Produces one workbook,
Automation_Test_Report.xlsx, with sheets matching the requested audit format.

"Skipped" is intentionally empty: this pipeline has no concept of a skipped
test — every scenario always executes for real, so there is nothing
legitimate to put there. It's included as an empty sheet rather than padded
with invented rows.
"""
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
PASS_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

ROW_FIELDS = ["TestID", "Category", "Module / Page", "Test Case", "Method",
              "Environment", "Status", "CWE", "OWASP", "Observed Result (evidence)", "Executed At"]


def load(path):
    if not os.path.exists(path):
        return None
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def autosize(ws):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in ws[letter])
        ws.column_dimensions[letter].width = min(max(12, max_len + 2), 80)


def write_rows_sheet(wb, title, rows, component_name=None):
    ws = wb.create_sheet(title)
    fields = [f for f in ROW_FIELDS if any(f in r for r in rows)] if rows else ROW_FIELDS
    if component_name and not any("Component" in r for r in rows):
        headers = ["Component"] + fields
    else:
        headers = (["Component"] if any("Component" in r for r in rows) else []) + fields
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for r in rows:
        row_values = []
        for h in headers:
            if h == "Component" and "Component" not in r:
                row_values.append(component_name or "")
            else:
                row_values.append(r.get(h, ""))
        ws.append(row_values)
        if r.get("Status") == "Pass":
            fill = PASS_FILL
        elif r.get("Status") == "Fail":
            fill = FAIL_FILL
        else:
            fill = None
        if fill:
            status_col = headers.index("Status") + 1 if "Status" in headers else None
            if status_col:
                ws.cell(row=ws.max_row, column=status_col).fill = fill
    ws.freeze_panes = "A2"
    autosize(ws)
    return ws


def generate(reports_dir, output_path):
    suites = {
        "Backend & Security": load(os.path.join(reports_dir, "backend_security_results.json")),
        "API Load Testing": load(os.path.join(reports_dir, "api_load_test_results.json")),
        "Website E2E": load(os.path.join(reports_dir, "web_e2e_results.json")),
        "Mobile App E2E": load(os.path.join(reports_dir, "mobile_e2e_results.json")),
    }

    all_rows = []
    for name, data in suites.items():
        if not data:
            continue
        for r in data["results"]:
            row = dict(r)
            row["Component"] = name
            all_rows.append(row)

    passed_rows = [r for r in all_rows if r.get("Status") == "Pass"]
    failed_rows = [r for r in all_rows if r.get("Status") == "Fail"]

    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    write_rows_sheet(wb, "Executed Test Cases", all_rows)
    write_rows_sheet(wb, "Passed Tests", passed_rows)
    write_rows_sheet(wb, "Failed Tests", failed_rows)
    ws_skipped = wb.create_sheet("Skipped Tests")
    ws_skipped.append(["Note"])
    ws_skipped["A1"].fill = HEADER_FILL
    ws_skipped["A1"].font = HEADER_FONT
    ws_skipped.append(["No skipped tests — every scenario in this pipeline always executes for real; "
                        "nothing is conditionally skipped, so there is nothing genuine to list here."])
    ws_skipped["A2"].alignment = Alignment(wrap_text=True)
    ws_skipped.column_dimensions["A"].width = 100

    # ---- Execution Metrics ----
    ws_metrics = wb.create_sheet("Execution Metrics")
    ws_metrics.append(["Component", "Total", "Passed", "Failed", "Pass Rate"])
    for cell in ws_metrics[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    grand_total = grand_passed = grand_failed = 0
    for name, data in suites.items():
        if not data:
            ws_metrics.append([name, "not run", "-", "-", "-"])
            continue
        ws_metrics.append([name, data["total"], data["passed"], data["failed"], f"{data['pass_rate']:.1f}%"])
        grand_total += data["total"]
        grand_passed += data["passed"]
        grand_failed += data["failed"]
    ws_metrics.append([])
    overall_rate = (grand_passed / grand_total * 100) if grand_total else 0.0
    ws_metrics.append(["TOTAL", grand_total, grand_passed, grand_failed, f"{overall_rate:.1f}%"])
    for cell in ws_metrics[ws_metrics.max_row]:
        cell.font = Font(bold=True)
    autosize(ws_metrics)

    # ---- Defect / Risk Summary (Failed rows, grouped, with CWE/OWASP if present) ----
    ws_defects = wb.create_sheet("Defect Summary")
    headers = ["Component", "TestID", "Module / Page", "Test Case", "CWE", "OWASP", "Evidence"]
    ws_defects.append(headers)
    for cell in ws_defects[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for r in failed_rows:
        ws_defects.append([
            r.get("Component", ""), r.get("TestID", ""), r.get("Module / Page", ""),
            r.get("Test Case", ""), r.get("CWE", ""), r.get("OWASP", ""),
            r.get("Observed Result (evidence)", ""),
        ])
    if not failed_rows:
        ws_defects.append(["No failed tests in this run.", "", "", "", "", "", ""])
    autosize(ws_defects)

    # ---- Pass Rate Summary (pie-chart-friendly single table) ----
    ws_rate = wb.create_sheet("Pass Rate Summary")
    ws_rate.append(["Status", "Count"])
    for cell in ws_rate[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws_rate.append(["Passed", grand_passed])
    ws_rate.append(["Failed", grand_failed])
    autosize(ws_rate)

    os.makedirs(os.path.dirname(output_path), exist_ok=True) if os.path.dirname(output_path) else None
    wb.save(output_path)
    return output_path, grand_total, grand_passed, grand_failed


if __name__ == "__main__":
    reports_dir = sys.argv[1] if len(sys.argv) > 1 else os.path.join(REPO_ROOT, "reports")
    output_path = sys.argv[2] if len(sys.argv) > 2 else os.path.join(REPO_ROOT, "reports", "Automation_Test_Report.xlsx")
    path, total, passed, failed = generate(reports_dir, output_path)
    print(f"Generated {path}")
    print(f"Total: {total}, Passed: {passed}, Failed: {failed}")
